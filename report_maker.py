import pandas as pd
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

def parse_ocr_data(text_data):
    """Parse the OCR analysis text data and extract structured information"""
    
    # Split data by images
    image_sections = text_data.split('--- Image')[1:]  # Skip the first empty split
    
    all_results = []
    summary_data = []
    
    for section in image_sections:
        lines = section.strip().split('\n')
        
        # Extract image info
        image_line = lines[0].split(':')[1].strip().replace(' ---', '')
        image_id = int(lines[0].split(':')[0])
        
        # Extract final result
        final_result = None
        method1_result = None
        method2_result = None
        
        for line in lines:
            if line.strip().startswith('✅ Best guess:'):
                final_result = re.search(r'(\d{10})', line)
                final_result = final_result.group(1) if final_result else None
            elif line.strip().startswith('Method 1 (Improved):'):
                method1_result = re.search(r'(\d{10})', line)
                method1_result = method1_result.group(1) if method1_result else None
            elif line.strip().startswith('Method 2 (Region):'):
                if 'None' in line:
                    method2_result = None
                else:
                    method2_result = re.search(r'(\d+)', line)
                    method2_result = method2_result.group(1) if method2_result else None
        
        # Extract individual OCR attempts
        ocr_methods = []
        candidates = []
        
        for line in lines:
            if line.strip().startswith('📄'):
                method_name = line.split(':')[0].replace('📄', '').strip()
                ocr_result = line.split(':')[1].strip()
                
                # Extract NID candidates from this line
                nid_matches = re.findall(r'\b(\d{10})\b', ocr_result)
                
                ocr_methods.append({
                    'image_id': image_id,
                    'image_file': image_line,
                    'method': method_name,
                    'raw_ocr_result': ocr_result,
                    'extracted_nids': ', '.join(nid_matches) if nid_matches else 'None',
                    'nid_count': len(nid_matches)
                })
                
                candidates.extend(nid_matches)
        
        all_results.extend(ocr_methods)
        
        # Count candidate frequencies for this image
        candidate_counter = Counter(candidates)
        
        summary_data.append({
            'image_id': image_id,
            'image_file': image_line,
            'total_ocr_methods': len(ocr_methods),
            'unique_candidates': len(candidate_counter),
            'most_common_candidate': candidate_counter.most_common(1)[0][0] if candidate_counter else None,
            'most_common_frequency': candidate_counter.most_common(1)[0][1] if candidate_counter else 0,
            'method1_result': method1_result,
            'method2_result': method2_result,
            'final_best_guess': final_result,
            'all_candidates': ', '.join([f"{k}({v})" for k, v in candidate_counter.most_common()])
        })
    
    return pd.DataFrame(all_results), pd.DataFrame(summary_data)

def create_excel_report(detailed_df, summary_df, image_paths=None, filename='OCR_Analysis_Report.xlsx'):
    """Create a formatted Excel report with multiple sheets"""
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write data to sheets
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
        
        # Get workbook and worksheets for formatting
        workbook = writer.book
        summary_ws = writer.sheets['Summary']
        detailed_ws = writer.sheets['Detailed Results']
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format Summary sheet
        for row in summary_ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # Auto-adjust column widths for summary
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format Detailed Results sheet
        for row in detailed_ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # Auto-adjust column widths for detailed results
        for column in detailed_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            detailed_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to all cells
        for ws in [summary_ws, detailed_ws]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        if image_paths:
            images_ws = workbook.create_sheet(title='Images')
            
            # Optionally write headers
            images_ws['A1'] = 'Image ID'
            images_ws['B1'] = 'Image Label'
            images_ws['C1'] = 'Image Preview'
            
            for idx, (label, img_path) in enumerate(image_paths, start=2):
                images_ws[f'A{idx}'] = idx-2
                images_ws[f'B{idx}'] = label
                
                img = XLImage(img_path)
                
                # Optionally resize images here (pixels) - openpyxl does not support easy resizing
                # But you can scale: img.width, img.height attributes to scale
                
                # Insert image in column C at row idx (for example, cell C2, C3, ...)
                cell_location = f'C{idx}'
                images_ws.add_image(img, cell_location)
                
                # Adjust row height for better image display
                images_ws.row_dimensions[idx].height = 100  # Adjust as needed
            
            # Adjust column widths to accommodate images & text
            images_ws.column_dimensions['A'].width = 10
            images_ws.column_dimensions['B'].width = 20
            images_ws.column_dimensions['C'].width = 40

    print(f"Excel report created successfully: {filename}")

def generate_analysis_stats(summary_df):
    """Generate additional analysis statistics"""
    
    stats = {
        'total_images_processed': len(summary_df),
        'images_with_consistent_results': len(summary_df[summary_df['method1_result'] == summary_df['final_best_guess']]),
        'average_unique_candidates_per_image': summary_df['unique_candidates'].mean(),
        'most_common_overall_result': summary_df['final_best_guess'].mode().iloc[0] if not summary_df['final_best_guess'].mode().empty else None,
        'method2_success_rate': (summary_df['method2_result'].notna().sum() / len(summary_df)) * 100
    }
    
    return stats

# Main execution
def main(image_paths_with_labels):
    # Your OCR data (paste your data here)
    ocr_text_data = """--- Image 0: nationalId_TC3.jpg ---
🚀 Method 1: Improved OCR

🔍 OCR Results for nationalId_TC3.jpg:
  📄 Adaptive+Whitelist digits: 490 01 5 9060
  📄 Adaptive+PSM 8 digits: 49001590609
  📄 Adaptive+PSM 7 digits: 490 01 5 9060
  📄 Adaptive+Default PSM 6: NID No. 490 01 5 9060 we ae
  📄 Adaptive+Single word: wow. 4900159060 9
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Whitelist digits: 44900159060
  📄 OTSU+PSM 8 digits: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  📄 OTSU+PSM 7 digits: 44900159060
  📄 OTSU+Default PSM 6: sons 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Single word: aon 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  📄 Simple+Whitelist digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+PSM 8 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+PSM 7 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Default PSM 6: ~. 490015 9060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Single word: ou. = 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  🎯 Best candidate: 4900159060

🎯 Method 2: Region-based OCR
  🎯 ROI OCR result: 404


📊 Final Results:
  Method 1 (Improved): 4900159060
  Method 2 (Region): None
  ✅ Best guess: 4900159060 (from Method 1)
------------------------------------------------------------

--- Image 1: nationalId_TC0.jpg ---
🚀 Method 1: Improved OCR

🔍 OCR Results for nationalId_TC0.jpg:
  📄 Adaptive+PSM 8 digits: 19001590602
  📄 Adaptive+Single word: Mp ter==190-015 9060. 2
  ✅ Found NID candidate: 1900159060
  📄 OTSU+Whitelist digits: 44900159060
  📄 OTSU+PSM 8 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+PSM 7 digits: 44900159060
  📄 OTSU+Default PSM 6: «ono 4900159060 —
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Single word: apne 4900159060 sy
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Whitelist digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+PSM 8 digits: 64900159060
  📄 Simple+PSM 7 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Default PSM 6: cose 490.015.9060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Single word: cone =©4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  🎯 Best candidate: 4900159060

🎯 Method 2: Region-based OCR
  🎯 ROI OCR result: 404


📊 Final Results:
  Method 1 (Improved): 4900159060
  Method 2 (Region): None
  ✅ Best guess: 4900159060 (from Method 1)
------------------------------------------------------------

--- Image 2: nationalId_TC2.jpg ---
🚀 Method 1: Improved OCR

🔍 OCR Results for nationalId_TC2.jpg:
  📄 Adaptive+Whitelist digits: 49001590608832
  📄 Adaptive+PSM 8 digits: 49001590608832
  📄 Adaptive+PSM 7 digits: 49001590608832
  📄 Adaptive+Default PSM 6: sono 4900159060 8832
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Adaptive+Single word: sono 4900159060 8832
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Whitelist digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+PSM 8 digits: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  📄 OTSU+PSM 7 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Default PSM 6: sono 4900159060 —
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Single word: sono §=6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  📄 Simple+Whitelist digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+PSM 8 digits: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  📄 Simple+PSM 7 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Default PSM 6: sono 4900159060 —
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Single word: sono §=6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  ✅ Found NID candidate: 6900459060
  🎯 Best candidate: 4900159060

🎯 Method 2: Region-based OCR
  🎯 ROI OCR result: 4004


📊 Final Results:
  Method 1 (Improved): 4900159060
  Method 2 (Region): None
  ✅ Best guess: 4900159060 (from Method 1)
------------------------------------------------------------

--- Image 3: nationalId_TC1.jpg ---
🚀 Method 1: Improved OCR

🔍 OCR Results for nationalId_TC1.jpg:
  📄 Adaptive+PSM 8 digits: 1900159060
  ✅ Found NID candidate: 1900159060
  ✅ Found NID candidate: 1900159060
  ✅ Found NID candidate: 1900159060
  📄 Adaptive+Single word: Mote ==190:015 9060. =
  ✅ Found NID candidate: 1900159060
  ✅ Found NID candidate: 1900159060
  📄 OTSU+Whitelist digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+PSM 8 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+PSM 7 digits: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Default PSM 6: «ono 4900159060 —
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 OTSU+Single word: non 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  ✅ Found NID candidate: 4900159060
  📄 Simple+Whitelist digits: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  📄 Simple+PSM 8 digits: 64900159060
  📄 Simple+PSM 7 digits: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  📄 Simple+Default PSM 6: cone =». 490.045 9060
  ✅ Found NID candidate: 4900459060
  ✅ Found NID candidate: 4900459060
  📄 Simple+Single word: cone =64900159060
  🎯 Best candidate: 4900159060

🎯 Method 2: Region-based OCR
  🎯 ROI OCR result: 404


📊 Final Results:
  Method 1 (Improved): 4900159060
  Method 2 (Region): None
  ✅ Best guess: 4900159060 (from Method 1)"""

    # Parse the data
    detailed_df, summary_df = parse_ocr_data(ocr_text_data)
    
    # Generate statistics
    stats = generate_analysis_stats(summary_df)
    
    # Create Excel report
    create_excel_report(detailed_df, summary_df, image_paths_with_labels)
    
    # Print summary statistics
    print("\n=== OCR Analysis Statistics ===")
    print(f"Total images processed: {stats['total_images_processed']}")
    print(f"Images with consistent results: {stats['images_with_consistent_results']}")
    print(f"Average unique candidates per image: {stats['average_unique_candidates_per_image']:.2f}")
    print(f"Most common overall result: {stats['most_common_overall_result']}")
    print(f"Method 2 success rate: {stats['method2_success_rate']:.1f}%")
    
    print(f"\nDetailed results shape: {detailed_df.shape}")
    print(f"Summary results shape: {summary_df.shape}")

if __name__ == "__main__":
    # ls history/6e62cbac-876d-4af1-97ab-95f8e42b507b/FrontSide/
# nationalId_TC0.jpg  nationalId_TC1.jpg  nationalId_TC2.jpg  nationalId_TC3.jpg
    temp = 'history/6e62cbac-876d-4af1-97ab-95f8e42b507b/FrontSide/nationalId_TC'

    keys = ['gray', 'resizedCopyImage', 'grayNew', 'enhanced']
    image_paths = [f"{temp}{i}.jpg" for i in range(4)]
    image_paths_with_labels = list(zip(keys, image_paths))
    print(f"Image paths: {image_paths_with_labels}")
    main(image_paths_with_labels)